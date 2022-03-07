#coding=utf8

from __future__ import division
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import easygui

import openpyxl


def nufftfreqs(M, df=1):
    """Compute the frequency range used in nufft for M frequency bins"""
    return df * np.arange(-(M // 2), M - (M // 2))


def nudft(x, y, M, df=1.0, iflag=1):
    """Non-Uniform Direct Fourier Transform"""
    sign = -1 if iflag < 0 else 1
    return (1 / len(x)) * np.dot(y, np.exp(sign * 1j * 2*np.pi * nufftfreqs(M, df) * x[:, np.newaxis]))
    
def nudft2(x, y, M, df=1.0, iflag=1):
    return nudft(x,y,M,df,iflag)**2 *df
    
file_path = easygui.fileopenbox(filetypes = [['*.csv',"CSV file"]])
if(file_path == None):
    exit()

csv = pd.read_csv(file_path, skiprows=4, sep=",")
head = csv.columns
row = csv.index

title = u'Спектр'

nums_head = easygui.multchoicebox(u'Выберите столбец для обработки. В случае суммы - выберете несколько',title,head[1:])
if(nums_head == None):
    exit()

fft = easygui.buttonbox(u'Выберете действие',u'Спектр',[u'График',u'Спектр'])

if fft==u'Спектр':
    msg = u'Введите данные'
    fieldName = [u'Начальный момент времени',u'Приблизительный шаг по времени',u'Количество пиковых частот',u'Название ряда']
    fieldValues = ['20','0.01','2',nums_head[0]]
    fieldValues = easygui.multenterbox(msg,title,fieldName,fieldValues)
    print(fieldValues[0])
    starttime = float(fieldValues[0])
    dt = float(fieldValues[1])
    min_fr = 0
    max_fr = 7
    num_pick = int(fieldValues[2])
    series = csv[csv[head[0]] > starttime][nums_head[0]]
    for i in range(1,len(nums_head)):
        series = series+csv[csv.Time > starttime][nums_head[i]]

    series = series - np.mean(series)
    print(series)
    time = csv[csv[head[0]]>starttime][head[0]]
    M = len(time)
    df = 1.0/M/dt

    freq = nufftfreqs(M,df)
    spectr = nudft(time,series,M,df)
    spectr = np.abs(spectr[int(M/2):])
    freq = freq[int(M/2):]
    print(max(spectr))
    #freq = np.fft.fftfreq(M,dt)
    #spectr = np.fft.rfft(series)
    #spectr = np.abs(spectr)
    #freq = freq[:M/2+1]
    
    pickst = np.r_[False, spectr[1:] > spectr[:-1]] & np.r_[spectr[:-1] > spectr[1:], False]
    picks = np.array([[i,spectr[i]] for i in range(len(freq)) if pickst[i]==True])
    
    maxfr = picks[picks[:,1].argsort()]
    maxfr = maxfr[-num_pick:]
    maxfr = maxfr[::-1]
    maxfr  = maxfr[:,0].astype(int)

    fig = plt.figure(figsize=(15,8))
    sizeOfFont = 12
    fontProperties = {'weight' : 'bold', 'size' : sizeOfFont}
    spec = fig.add_subplot(111)
    line, = spec.plot(freq,spectr,'b',linewidth=2,antialiased=True)

    spec.set_xlabel(u'Frequency, Hz', fontProperties)
    spec.set_title('Spectrum: "' + fieldValues[3] + '"')
    spec.set_xticks(np.arange(0, max_fr + 0.2, 0.5))
    spec.set_xticklabels(spec.get_xticks())
    spec.set_yticklabels(spec.get_yticks())
    spec.set_xlim((0, max_fr))
    spec.grid(True)

    for i in range(num_pick):
        maxpow = spectr[maxfr[i]]
        spec.annotate('%.3f' % freq[maxfr[i]] + ' Hz',
                      xy=(freq[maxfr[i]], maxpow), xycoords='data',
                      xytext=(freq[maxfr[i]], maxpow),
                      horizontalalignment='center', verticalalignment='bottom', fontsize=16)
    plt.show()
    fig.savefig('spectr.png')

    # wbip = openpyxl.Workbook()
    # wbip.create_sheet(title='Первый лист', index=0)
    # sheet = wbip['Первый лист']
    #
    # for i in range(1,len(spectr)+1):
    #     cell = sheet.cell(row=i, column=1)
    #     cell.value = spectr[i-1]
    #
    # wbip.save('example1.xlsx')
    # print(spectr,max(spectr))




if fft==u'График':
    msg = u'Введите данные'
    fieldName = [u'Начальный момент времени']
    fieldValues = ['10']
    fieldValues = easygui.multenterbox(msg,title,fieldName,fieldValues) 
    
    starttime = int(fieldValues[0])
    series = csv[csv.Time > starttime][nums_head[0]]
    for i in range(1,len(nums_head)):
        series = series+csv[csv.Time > starttime][nums_head[i]]
    time = csv[csv.Time>starttime][head[0]]
    
    fig = plt.figure(figsize=(15,8))
    
    sizeOfFont = 12
    fontProperties = {'weight' : 'bold', 'size' : sizeOfFont}
    avg = np.mean(series)
    minv = np.min(series)
    maxv = np.max(series)
    spec = fig.add_subplot(111)
    line, = spec.plot(time,series,linewidth=2,antialiased=True)
    spec.plot((np.min(time),np.max(time)),(minv,minv),'r--')
    spec.plot((np.min(time),np.max(time)),(maxv,maxv),'r--')
    spec.plot((np.min(time),np.max(time)),(avg,avg),'r--')
    
    spec.set_xlabel(u'Time, sec',fontProperties)
    spec.set_xticklabels(spec.get_xticks(), fontProperties)
    spec.set_yticklabels(spec.get_yticks(), fontProperties)
    spec.grid(True)
    
    spec.annotate('Min = %.3f' %minv,
    			xycoords='data',xy=(np.mean(time),minv), 
    			horizontalalignment='center', verticalalignment='bottom',fontsize=12,weight='bold',backgroundcolor='w')
    spec.annotate('Max = %.3f' %maxv,
    			xycoords='data',xy=(np.mean(time),maxv), 
    			horizontalalignment='center', verticalalignment='top',fontsize=12,weight='bold',backgroundcolor='w')
    spec.annotate('Avg = %.3f' %avg,
    			xycoords='data',xy=(np.mean(time),avg), 
    			horizontalalignment='center', verticalalignment='bottom',fontsize=12,weight='bold',backgroundcolor='w')
    plt.show()